"""Document Processing Service for Italian Tax Document Analysis.

Handles text extraction, structured data parsing, and document classification
for PDF, Excel, and CSV documents with Italian tax/financial content.
"""

import csv
import io
import locale
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import openpyxl
import pandas as pd
import pdfplumber
import PyPDF2
import pytesseract
from pdf2image import convert_from_bytes
from PIL import Image

from app.core.config import get_settings
from app.models.document_simple import ITALIAN_DOCUMENT_PATTERNS, Document, DocumentType, ItalianDocumentCategory


class DocumentProcessingError(Exception):
    """Exception raised during document processing"""

    pass


class DocumentProcessor:
    """Service for processing and analyzing Italian tax documents"""

    def __init__(self):
        self.settings = get_settings()
        # Set Italian locale for number parsing
        try:
            locale.setlocale(locale.LC_ALL, "it_IT.UTF-8")
        except locale.Error:
            try:
                locale.setlocale(locale.LC_ALL, "Italian_Italy.1252")  # Windows
            except locale.Error:
                pass  # Fallback to manual parsing

    async def extract_text(self, document: Document) -> dict[str, Any]:
        """Extract text from document using appropriate method.

        Args:
          document: Document model instance

        Returns:
          Dictionary with extraction results
        """
        try:
            if document.file_type == DocumentType.PDF.value:
                return await self._extract_pdf_text(document)
            elif document.file_type in [DocumentType.EXCEL_XLSX.value, DocumentType.EXCEL_XLS.value]:
                return await self._extract_excel_text(document)
            elif document.file_type == DocumentType.CSV.value:
                return await self._extract_csv_text(document)
            else:
                raise DocumentProcessingError(f"Unsupported file type: {document.file_type}")

        except Exception as e:
            return {
                "success": False,
                "error": str(e),
                "text": "",
                "text_length": 0,
                "page_count": 0,
                "ocr_used": False,
            }

    async def _extract_pdf_text(self, document: Document) -> dict[str, Any]:
        """Extract text from PDF using PyPDF2 and fallback to OCR"""
        # Read PDF file from storage
        file_path = self._get_document_storage_path(document)

        try:
            with open(file_path, "rb") as file:
                pdf_reader = PyPDF2.PdfReader(file)
                pages_text = []

                for _page_num, page in enumerate(pdf_reader.pages):
                    page_text = page.extract_text()
                    pages_text.append(page_text)

                full_text = "\n".join(pages_text)

                # If text extraction yields very little content, try OCR
                ocr_used = False
                if len(full_text.strip()) < 100:  # Likely scanned PDF
                    ocr_text = await self._extract_pdf_with_ocr(file_path)
                    if len(ocr_text) > len(full_text):
                        full_text = ocr_text
                        ocr_used = True

                return {
                    "success": True,
                    "text": full_text,
                    "text_length": len(full_text),
                    "page_count": len(pdf_reader.pages),
                    "ocr_used": ocr_used,
                }

        except Exception as e:
            raise DocumentProcessingError(f"PDF text extraction failed: {str(e)}")

    async def _extract_pdf_with_ocr(self, file_path: str) -> str:
        """Extract text from PDF using OCR (pytesseract)"""
        try:
            # Convert PDF pages to images
            images = convert_from_bytes(open(file_path, "rb").read())

            ocr_text_parts = []
            for image in images:
                # Use Italian language for OCR
                text = pytesseract.image_to_string(image, lang="ita+eng")
                ocr_text_parts.append(text)

            return "\n".join(ocr_text_parts)

        except Exception as e:
            raise DocumentProcessingError(f"OCR extraction failed: {str(e)}")

    async def _extract_excel_text(self, document: Document) -> dict[str, Any]:
        """Extract text content from Excel files"""
        file_path = self._get_document_storage_path(document)

        try:
            workbook = openpyxl.load_workbook(file_path)
            all_text = []

            for worksheet in workbook.worksheets:
                sheet_text = f"--- {worksheet.title} ---\n"

                for row in worksheet.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None:
                            row_text.append(str(cell))
                    if row_text:
                        sheet_text += " | ".join(row_text) + "\n"

                all_text.append(sheet_text)

            full_text = "\n".join(all_text)

            return {
                "success": True,
                "text": full_text,
                "text_length": len(full_text),
                "sheet_count": len(workbook.worksheets),
                "ocr_used": False,
            }

        except Exception as e:
            raise DocumentProcessingError(f"Excel text extraction failed: {str(e)}")

    async def _extract_csv_text(self, document: Document) -> dict[str, Any]:
        """Extract text content from CSV files"""
        file_path = self._get_document_storage_path(document)

        try:
            # Try multiple encodings for Italian support
            encodings = ["utf-8-sig", "utf-8", "iso-8859-1", "windows-1252"]

            for encoding in encodings:
                try:
                    with open(file_path, encoding=encoding) as file:
                        csv_reader = csv.reader(file)
                        rows = list(csv_reader)

                        text_lines = []
                        for row in rows:
                            text_lines.append(" | ".join(row))

                        full_text = "\n".join(text_lines)

                        return {
                            "success": True,
                            "text": full_text,
                            "text_length": len(full_text),
                            "row_count": len(rows),
                            "encoding_used": encoding,
                            "ocr_used": False,
                        }

                except UnicodeDecodeError:
                    continue

            raise DocumentProcessingError("Could not decode CSV with any supported encoding")

        except Exception as e:
            raise DocumentProcessingError(f"CSV text extraction failed: {str(e)}")

    async def extract_structured_data(self, document: Document) -> dict[str, Any]:
        """Extract structured data from document based on Italian tax document patterns.

        Args:
          document: Document with extracted text

        Returns:
          Dictionary with structured data extraction
        """
        if not document.extracted_text:
            return {"success": False, "error": "No extracted text available"}

        try:
            # Determine document type first
            doc_classification = await self.classify_document(document)
            doc_category = doc_classification.get("category")

            if doc_category == ItalianDocumentCategory.FATTURA_ELETTRONICA.value:
                return await self._extract_fattura_data(document.extracted_text)
            elif doc_category == ItalianDocumentCategory.F24.value:
                return await self._extract_f24_data(document.extracted_text)
            elif doc_category == ItalianDocumentCategory.BILANCIO.value:
                return await self._extract_bilancio_data(document)
            else:
                return await self._extract_generic_financial_data(document.extracted_text)

        except Exception as e:
            return {"success": False, "error": str(e)}

    async def _extract_fattura_data(self, text: str) -> dict[str, Any]:
        """Extract structured data from Fattura Elettronica"""
        patterns = {
            "numero": r"(?:numero|n\.?|fe)[\s:]*([\w\/\-]+)",
            "data": r"(?:data|del|date)[\s:]*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})",
            "partita_iva_cedente": r"(?:cedente|prestatore)[\s\S]{0,200}(?:p\.?iva|partita\s+iva)[\s:]*([A-Z]{2}\d{11}|\d{11})",
            "partita_iva_cessionario": r"(?:cessionario|committente)[\s\S]{0,200}(?:p\.?iva|partita\s+iva)[\s:]*([A-Z]{2}\d{11}|\d{11})",
            "imponibile": r"imponibile[\s:]*(?:€?\s*)?([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})",
            "iva": r"iva\s+(?:(\d+(?:,\d+)?)%?)[\s:]*(?:€?\s*)?([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})",
            "totale": r"totale[\s\S]{0,50}(?:€?\s*)?([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})",
        }

        extracted = {"document_type": "fattura_elettronica"}

        for field, pattern in patterns.items():
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
            if match:
                if field in ["imponibile", "totale"]:
                    # Parse Italian number format
                    extracted[field] = await self._parse_italian_number(match.group(1))
                elif field == "iva":
                    # Extract both percentage and amount
                    if len(match.groups()) > 1:
                        extracted["aliquota_iva"] = float(match.group(1).replace(",", ".")) if match.group(1) else None
                        extracted["iva"] = await self._parse_italian_number(match.group(2)) if match.group(2) else None
                elif field.startswith("partita_iva"):
                    entity = "cedente" if "cedente" in field else "cessionario"
                    if entity not in extracted:
                        extracted[entity] = {}
                    extracted[entity]["partita_iva"] = match.group(1)
                else:
                    extracted[field] = match.group(1)

        # Organize totals
        if any(k in extracted for k in ["imponibile", "iva", "totale"]):
            extracted["totali"] = {
                "imponibile": extracted.pop("imponibile", 0),
                "iva": extracted.pop("iva", 0),
                "totale": extracted.pop("totale", 0),
            }

        return {"success": True, "data": extracted}

    async def _extract_f24_data(self, text: str) -> dict[str, Any]:
        """Extract structured data from F24 tax payment form"""
        patterns = {
            "anno_riferimento": r"anno\s+(?:di\s+)?riferimento[\s:]*(\d{4})",
            "codice_fiscale": r"codice\s+fiscale[\s:]*([A-Z]{6}\d{2}[A-Z]\d{2}[A-Z]\d{3}[A-Z])",
            "tributi": r"codice\s+tributo[\s:]*(\d{4})[\s\S]{0,200}?([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})",
            "totale_versamenti": r"totale\s+versamenti[\s:]*(?:€?\s*)?([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})",
        }

        extracted = {"document_type": "f24"}

        # Extract basic information
        for field in ["anno_riferimento", "codice_fiscale", "totale_versamenti"]:
            pattern = patterns[field]
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                if field == "totale_versamenti":
                    extracted[field] = await self._parse_italian_number(match.group(1))
                else:
                    extracted[field] = match.group(1)

        # Extract tributi (tax payments)
        tributi_pattern = r"codice\s+tributo[\s:]*(\d{4})[\s\S]{0,100}?(?:([A-Z\s]+)[\s\S]{0,100}?)?([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})"
        tributi_matches = re.findall(tributi_pattern, text, re.IGNORECASE)

        tributi = []
        for match in tributi_matches:
            codice, descrizione, importo = match
            tributo = {"codice": codice, "importo": await self._parse_italian_number(importo)}
            if descrizione:
                tributo["descrizione"] = descrizione.strip()
            tributi.append(tributo)

        if tributi:
            extracted["tributi"] = tributi

        return {"success": True, "data": extracted}

    async def _extract_bilancio_data(self, document: Document) -> dict[str, Any]:
        """Extract structured data from Bilancio (balance sheet)"""
        if document.file_type in [DocumentType.EXCEL_XLSX.value, DocumentType.EXCEL_XLS.value]:
            return await self.process_excel(document)
        else:
            # For PDF bilanci, extract key financial figures
            text = document.extracted_text
            patterns = {
                "totale_attivo": r"totale\s+attivo[\s:]*(?:€?\s*)?([0-9]{1,3}(?:\.[0-9]{3})*(?:,[0-9]{2})?)",
                "totale_passivo": r"totale\s+passivo[\s:]*(?:€?\s*)?([0-9]{1,3}(?:\.[0-9]{3})*(?:,[0-9]{2})?)",
                "patrimonio_netto": r"patrimonio\s+netto[\s:]*(?:€?\s*)?([0-9]{1,3}(?:\.[0-9]{3})*(?:,[0-9]{2})?)",
                "utile_netto": r"utile\s+(?:netto|esercizio)[\s:]*(?:€?\s*)?([0-9]{1,3}(?:\.[0-9]{3})*(?:,[0-9]{2})?)",
                "ricavi": r"ricavi?[\s:]*(?:€?\s*)?([0-9]{1,3}(?:\.[0-9]{3})*(?:,[0-9]{2})?)",
            }

            extracted = {"document_type": "bilancio"}

            for field, pattern in patterns.items():
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    extracted[field] = await self._parse_italian_number(match.group(1))

            return {"success": True, "data": extracted}

    async def _extract_generic_financial_data(self, text: str) -> dict[str, Any]:
        """Extract generic financial data from any document"""
        # Extract all monetary amounts
        money_pattern = r"(?:€?\s*)?([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})"
        amounts = []

        for match in re.finditer(money_pattern, text):
            amount = await self._parse_italian_number(match.group(1))
            amounts.append(amount)

        # Extract dates
        date_pattern = r"(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})"
        dates = re.findall(date_pattern, text)

        return {
            "success": True,
            "data": {
                "document_type": "generic_financial",
                "monetary_amounts": amounts[:10],  # Limit to first 10
                "total_amounts_found": len(amounts),
                "dates_found": dates[:5],  # Limit to first 5
                "text_length": len(text),
            },
        }

    async def classify_document(self, document: Document) -> dict[str, Any]:
        """Classify Italian tax document based on content patterns.

        Args:
          document: Document with extracted text

        Returns:
          Dictionary with classification results
        """
        if not document.extracted_text:
            return {"category": ItalianDocumentCategory.OTHER.value, "confidence": 0}

        text = document.extracted_text.lower()
        max_confidence = 0
        best_category = ItalianDocumentCategory.OTHER

        for category, patterns in ITALIAN_DOCUMENT_PATTERNS.items():
            confidence = 0
            matches = 0

            for pattern in patterns:
                if re.search(pattern, text, re.IGNORECASE):
                    matches += 1

            # Calculate confidence based on pattern matches
            if matches > 0:
                confidence = min(95, 70 + (matches * 5))  # Base 70% + 5% per match, max 95%

                if confidence > max_confidence:
                    max_confidence = confidence
                    best_category = category

        return {"category": best_category.value, "confidence": max_confidence}

    async def process_excel(self, document: Document) -> dict[str, Any]:
        """Process Excel file and extract tabular data"""
        file_path = self._get_document_storage_path(document)

        try:
            workbook = openpyxl.load_workbook(file_path)
            tables = {}
            financial_data = {}

            for worksheet in workbook.worksheets:
                sheet_name = worksheet.title
                sheet_data = []

                for row in worksheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):  # Skip empty rows
                        sheet_data.append(list(row))

                if sheet_data:
                    tables[sheet_name] = sheet_data

                    # Extract financial indicators for Italian bilanci
                    if any(keyword in sheet_name.lower() for keyword in ["bilancio", "stato", "conto"]):
                        financial_indicators = await self._extract_excel_financial_indicators(sheet_data)
                        financial_data.update(financial_indicators)

            return {
                "success": True,
                "sheets_processed": len(tables),
                "tables": tables,
                "financial_analysis": financial_data,
            }

        except Exception as e:
            raise DocumentProcessingError(f"Excel processing failed: {str(e)}")

    async def _extract_excel_financial_indicators(self, sheet_data: list[list]) -> dict[str, float]:
        """Extract financial indicators from Excel sheet data"""
        indicators = {}

        # Common Italian balance sheet terms
        patterns = {
            "totale_attivo": ["totale attivo", "attivo totale"],
            "totale_passivo": ["totale passivo", "passivo totale"],
            "patrimonio_netto": ["patrimonio netto", "capitale netto"],
            "debiti": ["debiti totali", "debiti"],
            "ricavi": ["ricavi vendite", "ricavi", "vendite"],
            "utile_netto": ["utile netto", "risultato esercizio"],
            "ebitda": ["ebitda", "margine operativo lordo"],
            "ebit": ["ebit", "risultato operativo"],
        }

        for row in sheet_data:
            for col_idx, cell in enumerate(row):
                if isinstance(cell, str):
                    cell_lower = cell.lower()

                    for indicator, terms in patterns.items():
                        for term in terms:
                            if term in cell_lower:
                                # Look for numerical value in same row (next columns)
                                for val_idx in range(col_idx + 1, min(len(row), col_idx + 4)):
                                    if row[val_idx] is not None:
                                        try:
                                            if isinstance(row[val_idx], int | float):
                                                indicators[indicator] = float(row[val_idx])
                                            elif isinstance(row[val_idx], str):
                                                # Try to parse Italian number format
                                                parsed = await self._parse_italian_number(str(row[val_idx]))
                                                if parsed is not None:
                                                    indicators[indicator] = parsed
                                            break
                                        except (ValueError, TypeError):
                                            continue
                                break

        # Calculate additional ratios if base values are available
        if "debiti" in indicators and "patrimonio_netto" in indicators and indicators["patrimonio_netto"] != 0:
            indicators["leverage_ratio"] = indicators["debiti"] / indicators["patrimonio_netto"]

        return indicators

    async def process_csv(self, document: Document) -> dict[str, Any]:
        """Process CSV file with Italian number format support"""
        file_path = self._get_document_storage_path(document)

        try:
            # Try multiple encodings
            encodings = ["utf-8-sig", "utf-8", "iso-8859-1", "windows-1252"]

            for encoding in encodings:
                try:
                    df = pd.read_csv(file_path, encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise DocumentProcessingError("Could not decode CSV file")

            # Process Italian number formats
            numeric_columns = []
            for col in df.columns:
                if df[col].dtype == "object":
                    # Try to convert Italian numbers
                    converted = df[col].apply(self._parse_italian_number_safe)
                    if converted.notna().sum() > len(df) * 0.5:  # >50% convertible
                        df[col] = converted
                        numeric_columns.append(col)

            analysis = {}

            # IVA register analysis for Italian tax files
            if any(keyword in " ".join(df.columns).lower() for keyword in ["iva", "imposta", "aliquota"]):
                analysis = await self._analyze_iva_register(df)

            return {
                "success": True,
                "rows_processed": len(df),
                "columns": len(df.columns),
                "numeric_columns": numeric_columns,
                "data_preview": df.head(5).to_dict("records"),
                "iva_analysis": analysis,
            }

        except Exception as e:
            raise DocumentProcessingError(f"CSV processing failed: {str(e)}")

    async def _analyze_iva_register(self, df: pd.DataFrame) -> dict[str, Any]:
        """Analyze IVA register data"""
        analysis = {}

        # Find amount columns
        amount_cols = [
            col
            for col in df.columns
            if any(term in col.lower() for term in ["importo", "totale", "imponibile", "imposta"])
        ]

        if amount_cols:
            for col in amount_cols:
                if col in df.columns and df[col].dtype in ["float64", "int64"]:
                    total = df[col].sum()
                    analysis[f"totale_{col.lower()}"] = total

        # Analyze by VAT rate (aliquota)
        rate_cols = [col for col in df.columns if "aliquota" in col.lower()]
        if rate_cols and amount_cols:
            rate_col = rate_cols[0]
            amount_col = amount_cols[0]

            by_rate = df.groupby(rate_col)[amount_col].agg(["count", "sum"]).to_dict("index")
            analysis["by_aliquota"] = {
                str(k): {"count": v["count"], "imponibile": v["sum"]} for k, v in by_rate.items()
            }

        return analysis

    async def _parse_italian_number(self, value: str) -> float | None:
        """Parse Italian number format (1.234,56) to float"""
        if not isinstance(value, str):
            return None

        # Remove currency symbols and extra spaces
        cleaned = re.sub(r"[€$£¥]", "", value.strip())

        # Handle Italian number format: 1.234.567,89
        if "," in cleaned and cleaned.rfind(",") > cleaned.rfind("."):
            # Italian format: thousands separator (.), decimal separator (,)
            parts = cleaned.split(",")
            if len(parts) == 2:
                integer_part = parts[0].replace(".", "")
                decimal_part = parts[1]
                try:
                    return float(f"{integer_part}.{decimal_part}")
                except ValueError:
                    pass

        # Handle standard format: 1,234,567.89
        elif "." in cleaned and cleaned.rfind(".") > cleaned.rfind(","):
            # Standard format: thousands separator (,), decimal separator (.)
            try:
                return float(cleaned.replace(",", ""))
            except ValueError:
                pass

        # Try direct conversion
        try:
            return float(cleaned.replace(",", "."))
        except ValueError:
            return None

    def _parse_italian_number_safe(self, value) -> float | None:
        """Safe version of Italian number parsing for pandas apply"""
        try:
            import asyncio

            loop = asyncio.get_event_loop()
            return loop.run_until_complete(self._parse_italian_number(str(value))) if pd.notna(value) else None
        except:
            return None

    async def _detect_encoding(self, document: Document) -> dict[str, Any]:
        """Detect file encoding for text files"""
        file_path = self._get_document_storage_path(document)

        with open(file_path, "rb") as file:
            raw_data = file.read()

        # Check for BOM
        has_bom = raw_data.startswith(b"\xef\xbb\xbf")

        return {
            "encoding": "utf-8-sig" if has_bom else "utf-8",
            "has_bom": has_bom,
            "confidence": 0.95 if has_bom else 0.8,
        }

    def _get_document_storage_path(self, document: Document) -> str:
        """Get file system path for stored document"""
        # This would integrate with the secure storage system
        # For now, return a placeholder path
        storage_dir = Path("/tmp/document_storage")  # Would be configurable
        return str(storage_dir / document.filename)
