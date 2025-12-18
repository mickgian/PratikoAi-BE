"""TDD Tests for Document Processing Service Storage Integration.

DEV-007 Issue 5: Tests for SecureDocumentStorage integration fix.
These tests verify that DocumentProcessor correctly retrieves encrypted
document content from SecureDocumentStorage and processes it.
"""

import io
from unittest.mock import AsyncMock, MagicMock, patch
from uuid import uuid4

import pytest

from app.models.document_simple import Document, DocumentType, ProcessingStatus
from app.services.document_processing_service import DocumentProcessingError, DocumentProcessor


class TestDocumentProcessorStorageIntegration:
    """Test suite for SecureDocumentStorage integration (DEV-007 Issue 5)."""

    @pytest.fixture
    def processor(self):
        """Create a DocumentProcessor instance."""
        return DocumentProcessor()

    @pytest.fixture
    def sample_document(self):
        """Create a sample document for testing."""
        return Document(
            id=uuid4(),
            user_id=1,
            filename="test_file.xlsx",
            original_filename="fondo_pensione.xlsx",
            file_type=DocumentType.EXCEL_XLSX.value,
            file_size=1024,
            processing_status=ProcessingStatus.PROCESSING.value,
        )

    @pytest.fixture
    def sample_excel_content(self):
        """Create sample Excel file content as bytes."""
        # Create a minimal valid XLSX file in memory
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fondo Pensione"
        ws["A1"] = "Nome"
        ws["B1"] = "Contributo"
        ws["A2"] = "Mario Rossi"
        ws["B2"] = 1500.50
        ws["A3"] = "Luigi Verdi"
        ws["B3"] = 2000.00

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @pytest.fixture
    def sample_csv_content(self):
        """Create sample CSV content as bytes."""
        csv_text = """Nome,Contributo,Data
Mario Rossi,1500.50,2024-01-15
Luigi Verdi,2000.00,2024-02-20
Anna Bianchi,1750.25,2024-03-10"""
        return csv_text.encode("utf-8")

    @pytest.fixture
    def sample_pdf_content(self):
        """Create minimal PDF content as bytes."""
        # Minimal valid PDF
        return b"""%PDF-1.4
1 0 obj
<< /Type /Catalog /Pages 2 0 R >>
endobj
2 0 obj
<< /Type /Pages /Kids [3 0 R] /Count 1 >>
endobj
3 0 obj
<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Contents 4 0 R >>
endobj
4 0 obj
<< /Length 44 >>
stream
BT /F1 12 Tf 100 700 Td (Test Document) Tj ET
endstream
endobj
xref
0 5
0000000000 65535 f
0000000009 00000 n
0000000058 00000 n
0000000115 00000 n
0000000204 00000 n
trailer
<< /Size 5 /Root 1 0 R >>
startxref
296
%%EOF"""

    # ==========================================================================
    # Test: _get_document_content() method
    # ==========================================================================

    @pytest.mark.asyncio
    async def test_get_document_content_retrieves_from_secure_storage(
        self, processor, sample_document, sample_excel_content
    ):
        """Verify content is retrieved via SecureDocumentStorage.retrieve_document()."""
        # Mock SecureDocumentStorage
        mock_storage = MagicMock()
        mock_storage.retrieve_document = AsyncMock(
            return_value={
                "success": True,
                "content": sample_excel_content,
                "decrypted": True,
            }
        )
        processor.secure_storage = mock_storage

        # Call the method
        content = await processor._get_document_content(sample_document)

        # Verify
        mock_storage.retrieve_document.assert_called_once_with(sample_document.id)
        assert content == sample_excel_content
        assert isinstance(content, bytes)

    @pytest.mark.asyncio
    async def test_get_document_content_raises_error_on_storage_failure(self, processor, sample_document):
        """Verify error is raised when storage retrieval fails."""
        # Mock SecureDocumentStorage returning failure
        mock_storage = MagicMock()
        mock_storage.retrieve_document = AsyncMock(
            return_value={
                "success": False,
                "error": "Document not found in storage",
            }
        )
        processor.secure_storage = mock_storage

        # Call should raise error
        with pytest.raises(DocumentProcessingError) as exc_info:
            await processor._get_document_content(sample_document)

        assert "Failed to retrieve document" in str(exc_info.value)
        assert "Document not found in storage" in str(exc_info.value)

    # ==========================================================================
    # Test: Excel extraction from bytes
    # ==========================================================================

    @pytest.mark.asyncio
    async def test_extract_excel_text_from_bytes(self, processor, sample_document, sample_excel_content):
        """Verify Excel extraction works with in-memory bytes via BytesIO."""
        # Mock the secure storage to return our sample content
        mock_storage = MagicMock()
        mock_storage.retrieve_document = AsyncMock(
            return_value={
                "success": True,
                "content": sample_excel_content,
                "decrypted": True,
            }
        )
        processor.secure_storage = mock_storage

        # Call extraction
        result = await processor._extract_excel_text(sample_document)

        # Verify extraction succeeded
        assert result["success"] is True
        assert result["text_length"] > 0
        assert "Fondo Pensione" in result["text"]  # Sheet name
        assert "Mario Rossi" in result["text"]
        assert "1500.5" in result["text"] or "1500,5" in result["text"]
        assert result["sheet_count"] == 1
        assert result["ocr_used"] is False

    @pytest.mark.asyncio
    async def test_extract_excel_text_handles_multiple_sheets(self, processor):
        """Verify Excel extraction handles multiple worksheets."""
        # Create Excel with multiple sheets
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "Data1"

        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "Data2"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        content = buffer.getvalue()

        doc = Document(
            id=uuid4(),
            filename="multi.xlsx",
            file_type=DocumentType.EXCEL_XLSX.value,
        )

        mock_storage = MagicMock()
        mock_storage.retrieve_document = AsyncMock(return_value={"success": True, "content": content})
        processor.secure_storage = mock_storage

        result = await processor._extract_excel_text(doc)

        assert result["success"] is True
        assert result["sheet_count"] == 2
        assert "Sheet1" in result["text"]
        assert "Sheet2" in result["text"]
        assert "Data1" in result["text"]
        assert "Data2" in result["text"]

    # ==========================================================================
    # Test: CSV extraction from bytes
    # ==========================================================================

    @pytest.mark.asyncio
    async def test_extract_csv_text_from_bytes(self, processor, sample_csv_content):
        """Verify CSV extraction works with decoded bytes via StringIO."""
        doc = Document(
            id=uuid4(),
            filename="test.csv",
            file_type=DocumentType.CSV.value,
        )

        mock_storage = MagicMock()
        mock_storage.retrieve_document = AsyncMock(return_value={"success": True, "content": sample_csv_content})
        processor.secure_storage = mock_storage

        result = await processor._extract_csv_text(doc)

        assert result["success"] is True
        assert "Mario Rossi" in result["text"]
        assert "Luigi Verdi" in result["text"]
        assert "Anna Bianchi" in result["text"]
        assert result["row_count"] == 4  # Header + 3 data rows

    @pytest.mark.asyncio
    async def test_extract_csv_handles_italian_encoding(self, processor):
        """Verify CSV extraction handles Italian characters with various encodings."""
        # CSV with Italian characters
        csv_text = """Nome,Città,Importo
Giuseppe Verdi,Città della Pieve,1.234,56
François Müller,Düsseldorf,2.345,78"""

        doc = Document(
            id=uuid4(),
            filename="italian.csv",
            file_type=DocumentType.CSV.value,
        )

        # Test with ISO-8859-1 encoding (common for Italian)
        content = csv_text.encode("iso-8859-1")

        mock_storage = MagicMock()
        mock_storage.retrieve_document = AsyncMock(return_value={"success": True, "content": content})
        processor.secure_storage = mock_storage

        result = await processor._extract_csv_text(doc)

        assert result["success"] is True
        assert "Giuseppe Verdi" in result["text"]

    # ==========================================================================
    # Test: PDF extraction from bytes
    # ==========================================================================

    @pytest.mark.asyncio
    @pytest.mark.skipif(
        bool(__import__("os").environ.get("CI") or __import__("os").environ.get("GITHUB_ACTIONS")),
        reason="PDF OCR extraction requires poppler which is not installed in CI"
    )
    async def test_extract_pdf_text_from_bytes(self, processor, sample_pdf_content):
        """Verify PDF extraction works with in-memory bytes."""
        doc = Document(
            id=uuid4(),
            filename="test.pdf",
            file_type=DocumentType.PDF.value,
        )

        mock_storage = MagicMock()
        mock_storage.retrieve_document = AsyncMock(return_value={"success": True, "content": sample_pdf_content})
        processor.secure_storage = mock_storage

        result = await processor._extract_pdf_text(doc)

        # Basic PDF extraction should work (may have minimal/no text in our minimal PDF)
        assert result["success"] is True
        assert "page_count" in result
        assert result["page_count"] >= 1

    # ==========================================================================
    # Test: Error handling
    # ==========================================================================

    @pytest.mark.asyncio
    async def test_extract_text_returns_error_on_storage_failure(self, processor, sample_document):
        """Verify extract_text handles storage failures gracefully."""
        mock_storage = MagicMock()
        mock_storage.retrieve_document = AsyncMock(return_value={"success": False, "error": "Storage unavailable"})
        processor.secure_storage = mock_storage

        result = await processor.extract_text(sample_document)

        assert result["success"] is False
        assert "error" in result
        assert "Storage unavailable" in result["error"] or "Failed to retrieve" in result["error"]

    @pytest.mark.asyncio
    async def test_extract_text_returns_error_on_corrupted_content(self, processor, sample_document):
        """Verify extract_text handles corrupted file content gracefully."""
        # Return garbage bytes that aren't a valid Excel file
        mock_storage = MagicMock()
        mock_storage.retrieve_document = AsyncMock(return_value={"success": True, "content": b"not a valid xlsx file"})
        processor.secure_storage = mock_storage

        result = await processor.extract_text(sample_document)

        assert result["success"] is False
        assert "error" in result

    # ==========================================================================
    # Test: process_excel from bytes
    # ==========================================================================

    @pytest.mark.asyncio
    async def test_process_excel_from_bytes(self, processor, sample_document, sample_excel_content):
        """Verify process_excel works with in-memory bytes."""
        mock_storage = MagicMock()
        mock_storage.retrieve_document = AsyncMock(return_value={"success": True, "content": sample_excel_content})
        processor.secure_storage = mock_storage

        result = await processor.process_excel(sample_document)

        assert result["success"] is True
        assert result["sheets_processed"] == 1
        assert "tables" in result
        assert "Fondo Pensione" in result["tables"]

    # ==========================================================================
    # Test: Integration - full text extraction flow
    # ==========================================================================

    @pytest.mark.asyncio
    async def test_full_extraction_flow_excel(self, processor, sample_document, sample_excel_content):
        """Test complete extraction flow: storage -> bytes -> text."""
        mock_storage = MagicMock()
        mock_storage.retrieve_document = AsyncMock(return_value={"success": True, "content": sample_excel_content})
        processor.secure_storage = mock_storage

        # Call main extract_text entry point
        result = await processor.extract_text(sample_document)

        # Verify full flow succeeded
        assert result["success"] is True
        assert result["text_length"] > 0
        assert "Mario Rossi" in result["text"]
        assert "Luigi Verdi" in result["text"]

        # Verify storage was called
        mock_storage.retrieve_document.assert_called_once_with(sample_document.id)
