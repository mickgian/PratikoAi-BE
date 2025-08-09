#!/usr/bin/env python3
"""
Create sample Excel files for Italian tax document testing.

This script creates realistic Excel files that simulate Italian
tax documents for testing purposes.
"""

import os
from datetime import datetime, date
from decimal import Decimal
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def create_sample_f24_data():
    """Create sample F24 tax form data"""
    return {
        'Codice_Tributo': ['1001', '1040', '3844', '4033', '6781'],
        'Descrizione': [
            'IRPEF - Acconto Prima Rata',
            'IRPEF - Saldo',
            'IVA - Versamento Mensile',
            'Addizionale Regionale IRPEF',
            'Addizionale Comunale IRPEF'
        ],
        'Rateazione': ['', '', '', '', '01'],
        'Anno_Riferimento': [2024, 2023, 2024, 2023, 2023],
        'Importo_Debito': [2500.00, 1800.00, 3200.00, 350.00, 180.00],
        'Importo_Credito': [0.00, 0.00, 0.00, 0.00, 0.00],
        'Data_Scadenza': [
            '2024-06-17', '2024-06-17', '2024-02-16', 
            '2024-06-17', '2024-06-17'
        ]
    }

def create_sample_dichiarazione_redditi():
    """Create sample tax declaration data"""
    return {
        'Rigo': ['RN1', 'RN2', 'RN3', 'RN4', 'RN5', 'RN6', 'RN7', 'RN8'],
        'Descrizione': [
            'Redditi da lavoro dipendente',
            'Redditi da pensione',
            'Redditi da lavoro autonomo',
            'Redditi di impresa',
            'Redditi di capitale',
            'Redditi diversi',
            'Altri redditi',
            'Totale redditi'
        ],
        'Importo': [35000.00, 0.00, 12000.00, 0.00, 850.00, 200.00, 0.00, 48050.00],
        'Note': [
            'CUD 2024',
            '',
            'Consulenze professionali',
            '',
            'Interessi bancari',
            'Plusvalenze',
            '',
            'Reddito complessivo lordo'
        ]
    }

def create_sample_contabilita_iva():
    """Create sample VAT accounting data"""
    dates = pd.date_range('2024-01-01', '2024-12-31', freq='M')
    return {
        'Mese': [d.strftime('%m/%Y') for d in dates],
        'IVA_Vendite': [
            15420.00, 18650.00, 22180.00, 19850.00, 21200.00, 23800.00,
            25600.00, 24100.00, 22900.00, 26500.00, 28200.00, 30150.00
        ],
        'IVA_Acquisti': [
            8950.00, 10200.00, 11800.00, 10500.00, 12300.00, 13600.00,
            14200.00, 13100.00, 12800.00, 15200.00, 16800.00, 18500.00
        ],
        'IVA_Dovuta': [
            6470.00, 8450.00, 10380.00, 9350.00, 8900.00, 10200.00,
            11400.00, 11000.00, 10100.00, 11300.00, 11400.00, 11650.00
        ],
        'Liquidazione': [
            6470.00, 8450.00, 10380.00, 9350.00, 8900.00, 10200.00,
            11400.00, 11000.00, 10100.00, 11300.00, 11400.00, 11650.00
        ]
    }

def create_excel_with_multiple_sheets():
    """Create comprehensive Excel file with multiple sheets"""
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # 1. F24 Data Sheet
    ws_f24 = wb.create_sheet("F24_2024")
    f24_data = create_sample_f24_data()
    df_f24 = pd.DataFrame(f24_data)
    
    # Add headers
    ws_f24.append(['MODELLO F24 - ANNO 2024'])
    ws_f24.append(['Codice Fiscale: RSSMRA85M01H501Z'])
    ws_f24.append(['Cognome e Nome: ROSSI MARIO'])
    ws_f24.append([])  # Empty row
    
    # Add data
    for r in dataframe_to_rows(df_f24, index=False, header=True):
        ws_f24.append(r)
    
    # Style the F24 sheet
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    ws_f24['A1'].font = title_font
    
    # 2. Tax Declaration Sheet
    ws_dichiarazione = wb.create_sheet("Dichiarazione_Redditi")
    dichiarazione_data = create_sample_dichiarazione_redditi()
    df_dichiarazione = pd.DataFrame(dichiarazione_data)
    
    ws_dichiarazione.append(['DICHIARAZIONE DEI REDDITI 2024'])
    ws_dichiarazione.append(['Anno d\'imposta: 2023'])
    ws_dichiarazione.append([])
    
    for r in dataframe_to_rows(df_dichiarazione, index=False, header=True):
        ws_dichiarazione.append(r)
    
    # 3. VAT Accounting Sheet
    ws_iva = wb.create_sheet("Contabilita_IVA")
    iva_data = create_sample_contabilita_iva()
    df_iva = pd.DataFrame(iva_data)
    
    ws_iva.append(['CONTABILITÀ IVA - ANNO 2024'])
    ws_iva.append(['Regime IVA: Ordinario'])
    ws_iva.append([])
    
    for r in dataframe_to_rows(df_iva, index=False, header=True):
        ws_iva.append(r)
    
    # Apply some formatting
    for sheet in wb.worksheets:
        for cell in sheet[1]:  # First row
            cell.font = title_font
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    return wb

def create_simple_excel_files():
    """Create individual Excel files for different document types"""
    
    # Current directory
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 1. Simple F24 Excel
    df_f24 = pd.DataFrame(create_sample_f24_data())
    f24_path = os.path.join(current_dir, "sample_f24.xlsx")
    
    with pd.ExcelWriter(f24_path, engine='openpyxl') as writer:
        df_f24.to_excel(writer, sheet_name='F24', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['F24']
        
        # Add title
        worksheet.insert_rows(1, 3)
        worksheet['A1'] = 'MODELLO F24 - VERSAMENTI TRIBUTARI'
        worksheet['A2'] = f'Data creazione: {datetime.now().strftime("%d/%m/%Y")}'
        worksheet['A1'].font = Font(bold=True, size=14)
    
    # 2. VAT Register Excel
    df_iva = pd.DataFrame(create_sample_contabilita_iva())
    iva_path = os.path.join(current_dir, "sample_registro_iva.xlsx")
    
    with pd.ExcelWriter(iva_path, engine='openpyxl') as writer:
        df_iva.to_excel(writer, sheet_name='Registro_IVA', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Registro_IVA']
        
        worksheet.insert_rows(1, 3)
        worksheet['A1'] = 'REGISTRO IVA - LIQUIDAZIONI PERIODICHE'
        worksheet['A2'] = f'Anno di riferimento: 2024'
        worksheet['A1'].font = Font(bold=True, size=14)
    
    # 3. Comprehensive tax document
    wb_comprehensive = create_excel_with_multiple_sheets()
    comprehensive_path = os.path.join(current_dir, "sample_documenti_fiscali_completi.xlsx")
    wb_comprehensive.save(comprehensive_path)
    
    return [f24_path, iva_path, comprehensive_path]

if __name__ == "__main__":
    print("Creating sample Italian tax Excel documents...")
    
    try:
        created_files = create_simple_excel_files()
        print("✅ Successfully created Excel files:")
        for file_path in created_files:
            file_name = os.path.basename(file_path)
            print(f"  - {file_name}")
    
    except Exception as e:
        print(f"❌ Error creating Excel files: {e}")