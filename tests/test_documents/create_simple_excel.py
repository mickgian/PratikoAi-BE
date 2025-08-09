#!/usr/bin/env python3
"""
Create sample Excel files for Italian tax document testing using openpyxl.
"""

import os
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_f24_excel():
    """Create F24 tax form Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "F24_2024"
    
    # Title and header info
    ws['A1'] = 'MODELLO F24 - VERSAMENTI TRIBUTARI'
    ws['A2'] = 'Codice Fiscale: RSSMRA85M01H501Z'
    ws['A3'] = 'Cognome e Nome: ROSSI MARIO'
    ws['A4'] = f'Data creazione: {datetime.now().strftime("%d/%m/%Y")}'
    
    # Headers
    headers = [
        'Codice Tributo', 'Descrizione', 'Rateazione', 
        'Anno Riferimento', 'Importo Debito', 'Importo Credito', 'Data Scadenza'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data rows
    data = [
        ['1001', 'IRPEF - Acconto Prima Rata', '', 2024, 2500.00, 0.00, '17/06/2024'],
        ['1040', 'IRPEF - Saldo', '', 2023, 1800.00, 0.00, '17/06/2024'],
        ['3844', 'IVA - Versamento Mensile', '', 2024, 3200.00, 0.00, '16/02/2024'],
        ['4033', 'Addizionale Regionale IRPEF', '', 2023, 350.00, 0.00, '17/06/2024'],
        ['6781', 'Addizionale Comunale IRPEF', '01', 2023, 180.00, 0.00, '17/06/2024']
    ]
    
    for row_idx, row_data in enumerate(data, 7):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Style title
    ws['A1'].font = Font(bold=True, size=14)
    
    return wb

def create_iva_register_excel():
    """Create VAT register Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro_IVA"
    
    # Title
    ws['A1'] = 'REGISTRO IVA - LIQUIDAZIONI PERIODICHE'
    ws['A2'] = 'Anno di riferimento: 2024'
    ws['A3'] = 'Partita IVA: 01234567890'
    
    # Headers
    headers = ['Mese', 'IVA Vendite', 'IVA Acquisti', 'IVA Dovuta', 'Liquidazione']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Monthly data
    months = ['01/2024', '02/2024', '03/2024', '04/2024', '05/2024', '06/2024',
              '07/2024', '08/2024', '09/2024', '10/2024', '11/2024', '12/2024']
    
    iva_vendite = [15420.00, 18650.00, 22180.00, 19850.00, 21200.00, 23800.00,
                   25600.00, 24100.00, 22900.00, 26500.00, 28200.00, 30150.00]
    
    iva_acquisti = [8950.00, 10200.00, 11800.00, 10500.00, 12300.00, 13600.00,
                    14200.00, 13100.00, 12800.00, 15200.00, 16800.00, 18500.00]
    
    for row_idx, month in enumerate(months, 6):
        ws.cell(row=row_idx, column=1, value=month)
        ws.cell(row=row_idx, column=2, value=iva_vendite[row_idx - 6])
        ws.cell(row=row_idx, column=3, value=iva_acquisti[row_idx - 6])
        iva_dovuta = iva_vendite[row_idx - 6] - iva_acquisti[row_idx - 6]
        ws.cell(row=row_idx, column=4, value=iva_dovuta)
        ws.cell(row=row_idx, column=5, value=iva_dovuta)
    
    # Style title
    ws['A1'].font = Font(bold=True, size=14)
    
    return wb

def create_comprehensive_excel():
    """Create comprehensive tax document with multiple sheets"""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Sheet 1: F24
    ws1 = wb.create_sheet("F24_2024")
    ws1['A1'] = 'MODELLO F24 - VERSAMENTI TRIBUTARI'
    ws1['A1'].font = Font(bold=True, size=14)
    
    headers = ['Codice Tributo', 'Descrizione', 'Importo']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
    
    f24_data = [
        ['1001', 'IRPEF Acconto', 2500.00],
        ['3844', 'IVA Mensile', 3200.00],
        ['4033', 'Add. Regionale', 350.00]
    ]
    
    for row_idx, row_data in enumerate(f24_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=value)
    
    # Sheet 2: Dichiarazione Redditi
    ws2 = wb.create_sheet("Dichiarazione_Redditi")
    ws2['A1'] = 'DICHIARAZIONE DEI REDDITI 2024'
    ws2['A2'] = 'Anno d\'imposta: 2023'
    ws2['A1'].font = Font(bold=True, size=14)
    
    headers = ['Rigo', 'Descrizione', 'Importo']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
    
    dichiarazione_data = [
        ['RN1', 'Redditi da lavoro dipendente', 35000.00],
        ['RN3', 'Redditi da lavoro autonomo', 12000.00],
        ['RN5', 'Redditi di capitale', 850.00],
        ['RN8', 'Totale redditi', 47850.00]
    ]
    
    for row_idx, row_data in enumerate(dichiarazione_data, 5):
        for col_idx, value in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value)
    
    # Sheet 3: Bilancio
    ws3 = wb.create_sheet("Bilancio")
    ws3['A1'] = 'STATO PATRIMONIALE'
    ws3['A1'].font = Font(bold=True, size=14)
    
    headers = ['Voce', 'Attivo', 'Passivo']
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
    
    bilancio_data = [
        ['Immobilizzazioni', 150000.00, ''],
        ['Crediti', 45000.00, ''],
        ['Disponibilità liquide', 18000.00, ''],
        ['Capitale sociale', '', 100000.00],
        ['Riserve', '', 50000.00],
        ['Debiti', '', 63000.00]
    ]
    
    for row_idx, row_data in enumerate(bilancio_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            ws3.cell(row=row_idx, column=col_idx, value=value)
    
    return wb

def main():
    """Create all sample Excel files"""
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    files_created = []
    
    try:
        # F24 Excel
        wb_f24 = create_f24_excel()
        f24_path = os.path.join(current_dir, "sample_f24.xlsx")
        wb_f24.save(f24_path)
        files_created.append("sample_f24.xlsx")
        
        # VAT Register Excel
        wb_iva = create_iva_register_excel()
        iva_path = os.path.join(current_dir, "sample_registro_iva.xlsx")
        wb_iva.save(iva_path)
        files_created.append("sample_registro_iva.xlsx")
        
        # Comprehensive Excel
        wb_comprehensive = create_comprehensive_excel()
        comprehensive_path = os.path.join(current_dir, "sample_documenti_fiscali_completi.xlsx")
        wb_comprehensive.save(comprehensive_path)
        files_created.append("sample_documenti_fiscali_completi.xlsx")
        
        print("✅ Successfully created Excel files:")
        for filename in files_created:
            print(f"  - {filename}")
            
    except Exception as e:
        print(f"❌ Error creating Excel files: {e}")
        return False
    
    return True

if __name__ == "__main__":
    print("Creating sample Italian tax Excel documents...")
    main()